#!/usr/bin/env python3
"""
xlsx-tool — OPAL 스킬 공용 xlsx 읽기/쓰기 CLI 도구

Usage:
  xlsx-tool.py info   <file>
  xlsx-tool.py read   <file> [--sheet <name|index>] [--range <A1:Z100>] [--header-row <n>]
  xlsx-tool.py search <file> --keyword <text> [--sheet <name|index>] [--range <A1:Z100>]
  xlsx-tool.py write  <file> --data <json> [--sheet <name>] [--mode new|update] [--format]
                             [--data-file <path>]
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string


# ── Helpers ──────────────────────────────────────────────────────────────────

def ok(command, **kwargs):
    print(json.dumps({"ok": True, "command": command, **kwargs}, ensure_ascii=False, default=str))

def err(command, message):
    print(json.dumps({"ok": False, "command": command, "error": message}, ensure_ascii=False))
    sys.exit(1)

def parse_range(range_str):
    """'A1:C10' → (min_row, min_col, max_row, max_col)"""
    try:
        start, end = range_str.upper().split(":")
        sc, sr = coordinate_from_string(start)
        ec, er = coordinate_from_string(end)
        return int(sr), column_index_from_string(sc), int(er), column_index_from_string(ec)
    except Exception:
        return None

def load_wb(filepath, data_only=False):
    p = Path(filepath)
    if not p.exists():
        return None, f"파일을 찾을 수 없습니다: {filepath}"
    if p.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return None, f"지원하지 않는 파일 형식입니다: {p.suffix}"
    try:
        return openpyxl.load_workbook(filepath, data_only=data_only), None
    except Exception as e:
        return None, str(e)

def get_sheet(wb, sheet_arg):
    if sheet_arg is None:
        return wb.active, None
    if sheet_arg.isdigit():
        idx = int(sheet_arg)
        if idx < 0 or idx >= len(wb.sheetnames):
            return None, f"시트 인덱스 범위 초과: {idx} (전체 {len(wb.sheetnames)}개)"
        return wb[wb.sheetnames[idx]], None
    if sheet_arg in wb.sheetnames:
        return wb[sheet_arg], None
    return None, f"시트를 찾을 수 없습니다: '{sheet_arg}' (전체: {wb.sheetnames})"

def apply_format(ws, header_row=1):
    """헤더 볼드 + 열 너비 자동 + 테두리"""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                length = len(str(cell.value))
                col = cell.column
                col_widths[col] = max(col_widths.get(col, 0), length)

            if cell.row == header_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", start_color="D9E1F2")
                cell.alignment = Alignment(horizontal="center")

            cell.border = border

    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(width + 4, 60)


# ── Commands ─────────────────────────────────────────────────────────────────

def cmd_info(args):
    wb, error = load_wb(args.file)
    if error:
        err("info", error)

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers = []
        if ws.max_row >= 1:
            headers = [
                str(cell.value) if cell.value is not None else ""
                for cell in ws[1]
                if cell.value is not None
            ]
        sheets.append({
            "name": name,
            "rows": ws.max_row,
            "columns": ws.max_column,
            "headers": headers,
        })

    ok("info", file=args.file, sheets=sheets)


def cmd_read(args):
    wb, error = load_wb(args.file, data_only=True)
    if error:
        err("read", error)

    header_row = int(args.header_row) if args.header_row else 1

    target_sheets = wb.sheetnames if args.sheet is None else None
    if args.sheet is not None:
        ws, e = get_sheet(wb, args.sheet)
        if e:
            err("read", e)
        target_sheets = [ws.title]

    result = {}
    for sheet_name in target_sheets:
        ws = wb[sheet_name]

        if args.range:
            bounds = parse_range(args.range)
            if not bounds:
                err("read", f"범위 형식 오류: {args.range} (예: A1:Z100)")
            min_row, min_col, max_row, max_col = bounds
            rows = list(ws.iter_rows(
                min_row=min_row, max_row=max_row,
                min_col=min_col, max_col=max_col,
                values_only=True
            ))
            result[sheet_name] = [list(r) for r in rows]
        else:
            headers = None
            data = []
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i == header_row:
                    headers = [str(v) if v is not None else f"col_{j}" for j, v in enumerate(row)]
                elif headers is not None:
                    data.append(dict(zip(headers, row)))
            result[sheet_name] = data if headers else [list(r) for r in ws.iter_rows(values_only=True)]

    if args.sheet is not None:
        ok("read", file=args.file, sheet=args.sheet, data=result[list(result.keys())[0]])
    else:
        ok("read", file=args.file, data=result)


def cmd_search(args):
    if not args.keyword:
        err("search", "--keyword 옵션이 필요합니다")

    wb, error = load_wb(args.file, data_only=True)
    if error:
        err("search", error)

    target_sheets = wb.sheetnames if args.sheet is None else None
    if args.sheet is not None:
        ws, e = get_sheet(wb, args.sheet)
        if e:
            err("search", e)
        target_sheets = [ws.title]

    keyword = args.keyword.lower()
    matches = []

    for sheet_name in target_sheets:
        ws = wb[sheet_name]

        if args.range:
            bounds = parse_range(args.range)
            if not bounds:
                err("search", f"범위 형식 오류: {args.range}")
            min_row, min_col, max_row, max_col = bounds
            rows = ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
        else:
            rows = ws.iter_rows()

        for row in rows:
            for cell in row:
                if cell.value is not None and keyword in str(cell.value).lower():
                    matches.append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "value": cell.value,
                    })

    ok("search", file=args.file, keyword=args.keyword, count=len(matches), matches=matches)


def cmd_write(args):
    # 데이터 로드
    if args.data_file:
        try:
            with open(args.data_file, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            err("write", f"data-file 로드 실패: {e}")
    elif args.data:
        try:
            data = json.loads(args.data)
        except Exception as e:
            err("write", f"--data JSON 파싱 실패: {e}")
    else:
        err("write", "--data 또는 --data-file 중 하나가 필요합니다")

    sheet_name = args.sheet or "Sheet1"
    mode = args.mode or "new"
    filepath = args.file

    if mode == "update" and Path(filepath).exists():
        wb, error = load_wb(filepath)
        if error:
            err("write", error)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

    # 데이터 쓰기
    if isinstance(data, list):
        if len(data) == 0:
            err("write", "데이터가 비어 있습니다")

        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            if mode != "update" or ws.max_row == 1:
                ws.append(headers)
            for row in data:
                ws.append([row.get(h) for h in headers])
        elif isinstance(data[0], list):
            for row in data:
                ws.append(row)
        else:
            ws.append(data)
    else:
        err("write", "data는 list(dict) 또는 list(list) 형식이어야 합니다")

    if args.format:
        apply_format(ws)

    try:
        wb.save(filepath)
    except Exception as e:
        err("write", f"파일 저장 실패: {e}")

    ok("write", file=filepath, sheet=sheet_name, mode=mode, rows_written=len(data))


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(prog="xlsx-tool", add_help=True)
    sub = parser.add_subparsers(dest="command")

    # info
    p_info = sub.add_parser("info")
    p_info.add_argument("file")

    # read
    p_read = sub.add_parser("read")
    p_read.add_argument("file")
    p_read.add_argument("--sheet")
    p_read.add_argument("--range")
    p_read.add_argument("--header-row", default="1")

    # search
    p_search = sub.add_parser("search")
    p_search.add_argument("file")
    p_search.add_argument("--keyword", required=True)
    p_search.add_argument("--sheet")
    p_search.add_argument("--range")

    # write
    p_write = sub.add_parser("write")
    p_write.add_argument("file")
    p_write.add_argument("--sheet")
    p_write.add_argument("--mode", choices=["new", "update"], default="new")
    p_write.add_argument("--data")
    p_write.add_argument("--data-file")
    p_write.add_argument("--format", action="store_true")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    try:
        {"info": cmd_info, "read": cmd_read, "search": cmd_search, "write": cmd_write}[args.command](args)
    except SystemExit:
        raise
    except Exception as e:
        err(args.command, str(e))


if __name__ == "__main__":
    main()
